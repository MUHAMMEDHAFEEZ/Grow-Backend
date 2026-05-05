"""
dashboard/services.py — Business logic: risk scoring, insights, recommendations, exports.

All business rules live here. Delegates read queries to selectors.py.
"""

from datetime import timedelta
from io import BytesIO

from django.utils import timezone

from accounts.models import User
from courses.models import Course, Enrollment
from core.events import EventBus, Events
from django.template.loader import render_to_string

from .models import (
    DashboardInsight,
    InterventionRecord,
    RecommendedAction,
    StudentNote,
)
from .selectors import (
    calculate_risk_score,
    get_class_detail,
    get_gpa_trend_for_class,
    get_performance_trend_for_student,
    get_risk_summary,
)


def generate_insights() -> list[dict]:
    """Run insight rules and create DashboardInsight records.

    Rules:
    1. Overcrowded classes (enrollment > 110% of target_capacity)
    2. Performance drops (class GPA decline > 10% MoM)
    3. Teacher performance decline
    4. At-risk students (risk_score > 60)
    """
    created = []

    # Rule 1: Overcrowded classes
    for course in Course.objects.all():
        enrollment_count = Enrollment.objects.filter(course=course).count()
        if course.target_capacity > 0:
            utilization = (enrollment_count / course.target_capacity) * 100
            if utilization > 110:
                insight, _ = DashboardInsight.objects.get_or_create(
                    insight_type=DashboardInsight.InsightType.OVERCROWDED_CLASS,
                    related_object_type="course",
                    related_object_id=course.id,
                    is_dismissed=False,
                    defaults={
                        "title": f"{course.title} is overcrowded",
                        "description": (
                            f"Enrollment {enrollment_count} exceeds "
                            f"{course.target_capacity} capacity by {utilization:.0f}%"
                        ),
                        "severity": DashboardInsight.Severity.WARNING,
                        "recommendation": "Consider splitting the class or increasing capacity",
                    },
                )
                if insight._state.adding:
                    created.append(insight)
                    EventBus.publish(
                        Events.INSIGHT_GENERATED,
                        {
                            "insight_id": insight.id,
                            "insight_type": insight.insight_type,
                            "severity": insight.severity,
                            "related_object_type": insight.related_object_type,
                            "related_object_id": insight.related_object_id,
                        },
                    )

    # Rule 2: Performance drops
    for course in Course.objects.all():
        trend = get_gpa_trend_for_class(course.id, months=2)
        if len(trend) >= 2:
            prev = trend[-2]["class_avg"]
            curr = trend[-1]["class_avg"]
            if prev > 0:
                change_pct = ((curr - prev) / prev) * 100
                if change_pct < -10:
                    insight, _ = DashboardInsight.objects.get_or_create(
                        insight_type=DashboardInsight.InsightType.PERFORMANCE_DROP,
                        related_object_type="course",
                        related_object_id=course.id,
                        is_dismissed=False,
                        defaults={
                            "title": f"{course.title} performance declining",
                            "description": (
                                f"GPA dropped {abs(change_pct):.1f}% month-over-month "
                                f"({prev:.2f} → {curr:.2f})"
                            ),
                            "severity": DashboardInsight.Severity.CRITICAL,
                            "recommendation": "Review teaching methods and student feedback",
                        },
                    )
                    if insight._state.adding:
                        created.append(insight)

    # Rule 3: At-risk students
    students = User.objects.filter(role=User.Role.STUDENT)
    for student in students:
        risk = calculate_risk_score(student.id)
        if risk["score"] > 60:
            severity = (
                DashboardInsight.Severity.CRITICAL
                if risk["score"] > 80
                else DashboardInsight.Severity.WARNING
            )
            insight, _ = DashboardInsight.objects.get_or_create(
                insight_type=DashboardInsight.InsightType.AT_RISK_STUDENT,
                related_object_type="student",
                related_object_id=student.id,
                is_dismissed=False,
                defaults={
                    "title": f"Student {student.username} at risk",
                    "description": (
                        f"Risk score: {risk['score']:.0f} ({risk['tier']}). "
                        f"Factors: {', '.join(risk['factors'])}"
                    ),
                    "severity": severity,
                    "recommendation": "Schedule parent meeting and review intervention plan",
                },
            )
            if insight._state.adding:
                created.append(insight)

    return [
        {
            "id": i.id,
            "title": i.title,
            "severity": i.severity,
            "type": i.insight_type,
        }
        for i in created
    ]


def dismiss_insight(insight_id: int, user) -> DashboardInsight | None:
    """Mark an insight as dismissed with audit trail."""
    try:
        insight = DashboardInsight.objects.get(id=insight_id)
        insight.is_dismissed = True
        insight.dismissed_by = user
        insight.dismissed_at = timezone.now()
        insight.save()
        return insight
    except DashboardInsight.DoesNotExist:
        return None


def add_student_note(
    student_id: int, author_id: int, note_text: str
) -> StudentNote | None:
    """Create a timestamped note on a student profile."""
    if not note_text or not note_text.strip():
        return None

    student = User.objects.filter(id=student_id, role=User.Role.STUDENT).first()
    author = User.objects.filter(id=author_id).first()
    if not student or not author:
        return None

    return StudentNote.objects.create(
        student=student,
        author=author,
        note=note_text.strip(),
    )


def create_intervention(
    student_id: int,
    action: str,
    priority: int = 1,
    assigned_to_id: int | None = None,
) -> InterventionRecord | None:
    """Create an intervention record for an at-risk student."""
    student = User.objects.filter(id=student_id, role=User.Role.STUDENT).first()
    if not student:
        return None

    assigned_to = None
    if assigned_to_id:
        assigned_to = User.objects.filter(id=assigned_to_id).first()

    return InterventionRecord.objects.create(
        student=student,
        action=action,
        priority=priority,
        assigned_to=assigned_to,
    )


def generate_recommendations(student_id: int) -> list[dict]:
    """Return prioritized recommended actions for an at-risk student."""
    risk = calculate_risk_score(student_id)
    actions = []

    if "low_gpa" in risk["factors"]:
        actions.append({
            "action": RecommendedAction.TUTORING_REFERRAL,
            "priority": 1,
            "reason": "Low GPA detected",
        })
    if "high_absences" in risk["factors"]:
        actions.append({
            "action": RecommendedAction.PARENT_MEETING,
            "priority": 1,
            "reason": "High absence rate",
        })
    if "disciplinary_cases" in risk["factors"]:
        actions.append({
            "action": RecommendedAction.COUNSELING_REFERRAL,
            "priority": 2,
            "reason": "Disciplinary cases on record",
        })
    if "declining_trend" in risk["factors"]:
        actions.append({
            "action": RecommendedAction.TEACHER_CHECKIN,
            "priority": 2,
            "reason": "Declining performance trend",
        })

    if not actions:
        actions.append({
            "action": RecommendedAction.PEER_MENTORING,
            "priority": 3,
            "reason": "Preventive support recommended",
        })

    return actions[:3]


def generate_report_pdf(filters: dict) -> bytes:
    """Render report HTML template to PDF using WeasyPrint."""
    from weasyprint import HTML

    report_data = get_risk_summary()
    context = {
        "filters": filters,
        "report": report_data,
        "generated_at": timezone.now().isoformat(),
    }
    html = render_to_string("dashboard/report_pdf.html", context)
    pdf = HTML(string=html).write_pdf()
    return pdf


def generate_report_excel(filters: dict) -> BytesIO:
    """Create structured Excel workbook using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["School Management Dashboard Report"])
    summary_sheet.append([f"Generated: {timezone.now().isoformat()}"])
    summary_sheet.append([f"Filters: {filters}"])
    summary_sheet.append([])

    risk = get_risk_summary()
    summary_sheet.append(["Risk Summary"])
    summary_sheet.append(["Tier", "Count"])
    for tier, count in risk["by_tier"].items():
        summary_sheet.append([tier.title(), count])

    data_sheet = wb.create_sheet("Data")
    data_sheet.append(["Student ID", "Name", "Risk Score", "Risk Tier", "Factors"])
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for s in risk.get("top_at_risk_students", []):
        data_sheet.append([
            s["student_id"],
            s["name"],
            s["risk_score"],
            s["risk_tier"],
            ", ".join(s["contributing_factors"]),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
